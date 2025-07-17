

import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# --- Utility Functions ---
def strip_leading_zeros(df):
    return df.applymap(lambda x: x.lstrip('0') if isinstance(x, str) and x.isdigit() else x)

def get_composite_key(df, key_columns):
    return df[key_columns].astype(str).apply(lambda x: x.str.strip()).fillna('NA').agg('-'.join, axis=1)

def compare_data(df1, df2, key_columns):
    df1 = strip_leading_zeros(df1).astype(str)
    df2 = strip_leading_zeros(df2).astype(str)
    df1['composite_key'] = get_composite_key(df1, key_columns)
    df2['composite_key'] = get_composite_key(df2, key_columns)
    if df1['composite_key'].duplicated().any():
        st.error("Duplicate composite keys found in first file.")
        return None, None, None, None, None
    if df2['composite_key'].duplicated().any():
        st.error("Duplicate composite keys found in second file.")
        return None, None, None, None, None
    df1.set_index('composite_key', inplace=True)
    df2.set_index('composite_key', inplace=True)
    added = df2.loc[~df2.index.isin(df1.index)]
    removed = df1.loc[~df1.index.isin(df2.index)]
    changed = [pd.concat([df1.loc[[key]], df2.loc[[key]]], keys=['file1', 'file2'])
               for key in df1.index.intersection(df2.index)
               if not df1.loc[key].equals(df2.loc[key])]
    changed_df = pd.concat(changed) if changed else pd.DataFrame()
    columns_added = [col for col in df2.columns if col not in df1.columns]
    columns_removed = [col for col in df1.columns if col not in df2.columns]
    return added, removed, changed_df, columns_added, columns_removed

def compare_sql_files(sql1, sql2):
    import difflib
    sql1_lines = [line.rstrip('\n') for line in sql1.getvalue().decode().splitlines()]
    sql2_lines = [line.rstrip('\n') for line in sql2.getvalue().decode().splitlines()]
    diff = list(difflib.ndiff(sql1_lines, sql2_lines))
    overlay_rows = []
    line_num1 = 1
    line_num2 = 1
    for line in diff:
        if line.startswith('+ '):
            overlay_rows.append({'Line': line[2:], 'Status': 'Added', 'LineNum1': '', 'LineNum2': line_num2})
            line_num2 += 1
        elif line.startswith('- '):
            overlay_rows.append({'Line': line[2:], 'Status': 'Removed', 'LineNum1': line_num1, 'LineNum2': ''})
            line_num1 += 1
        elif line.startswith('  '):
            overlay_rows.append({'Line': line[2:], 'Status': 'Unchanged', 'LineNum1': line_num1, 'LineNum2': line_num2})
            line_num1 += 1
            line_num2 += 1
    return pd.DataFrame(overlay_rows)

def suggest_key_columns(df1, df2):
    # Only consider columns present in both files
    common_cols = [col for col in df1.columns if col in df2.columns]
    # Prefer columns named 'id' or ending with '_id' that are unique in both files
    id_like = [col for col in common_cols if col.lower() == 'id' or col.lower().endswith('_id')]
    for col in id_like:
        if df1[col].is_unique and df2[col].is_unique:
            return [col]
    # Any unique column in both files
    for col in common_cols:
        if df1[col].is_unique and df2[col].is_unique:
            return [col]
    # Try all combinations from 2 up to all common columns
    from itertools import combinations
    n_cols = len(common_cols)
    for r in range(2, n_cols + 1):
        for combo in combinations(common_cols, r):
            if (
                df1[list(combo)].drop_duplicates().shape[0] == df1.shape[0]
                and df2[list(combo)].drop_duplicates().shape[0] == df2.shape[0]
            ):
                return list(combo)
    # Fallback: all common columns
    return list(common_cols)

# --- Streamlit UI ---
st.title("CSV & SQL Comparison Tool")

st.subheader("1. Upload CSV Files")
csv1 = st.file_uploader("Upload first CSV file", type="csv", key="csv1")
csv2 = st.file_uploader("Upload second CSV file", type="csv", key="csv2")

df1 = df2 = None
key_columns = []

if csv1 is not None and csv2 is not None:
    try:
        df1 = pd.read_csv(csv1)
        df2 = pd.read_csv(csv2)
        st.success("CSV files loaded.")
        st.subheader("2. Specify columns for a unique key (auto-suggested)")
        suggested_keys = suggest_key_columns(df1, df2)
        key_columns = st.multiselect(
            "Select key columns (must uniquely identify rows)",
            options=list(df1.columns),
            default=suggested_keys
        )
    except Exception as e:
        st.error(f"Error loading CSVs: {e}")

st.subheader("3. Upload SQL Files (optional)")
sql1 = st.file_uploader("Upload first SQL file", type="sql", key="sql1")
sql2 = st.file_uploader("Upload second SQL file", type="sql")

if st.button("Compare"):
    if df1 is not None and df2 is not None and key_columns:
        progress = st.progress(0, text="Starting comparison...")
        added, removed, changed, columns_added, columns_removed = compare_data(df1, df2, key_columns)
        progress.progress(40, text="Compared CSV data.")
        sql_overlay_df = None
        if sql1 and sql2:
            sql_overlay_df = compare_sql_files(sql1, sql2)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if added is not None and not added.empty:
                added.to_excel(writer, sheet_name='Added')
            if removed is not None and not removed.empty:
                removed.to_excel(writer, sheet_name='Removed')
            if changed is not None and not changed.empty:
                changed.to_excel(writer, sheet_name='Changed')
            if columns_added:
                pd.DataFrame({'Columns Added': columns_added}).to_excel(writer, sheet_name='Columns_Added', index=False)
            if columns_removed:
                pd.DataFrame({'Columns Removed': columns_removed}).to_excel(writer, sheet_name='Columns_Removed', index=False)
            if sql_overlay_df is not None and not sql_overlay_df.empty:
                sql_overlay_df.to_excel(writer, sheet_name='SQL_Overlay', index=False)
        progress.progress(70, text="Wrote results to Excel.")

        # Apply coloring to SQL overlay if present
        if sql_overlay_df is not None and not sql_overlay_df.empty:
            output.seek(0)
            wb = load_workbook(output)
            ws = wb['SQL_Overlay']
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                status = row[1].value
                if status == 'Added':
                    row[0].font = Font(color='006100')
                    row[0].fill = green_fill
                elif status == 'Removed':
                    row[0].font = Font(color='9C0006')
                    row[0].fill = red_fill
            # Save the colored workbook back to output
            temp_output = io.BytesIO()
            wb.save(temp_output)
            output = temp_output

        progress.progress(90, text="Finalizing...")
        st.success("Comparison complete!")
        progress.progress(100, text="Done!")

        st.download_button("Download comparison_results.xlsx", data=output.getvalue(), file_name="comparison_results.xlsx")

        if added is not None and not added.empty:
            st.write(f"### Added Rows ({len(added)})", added)
        if removed is not None and not removed.empty:
            st.write(f"### Removed Rows ({len(removed)})", removed)
        if columns_added:
            st.write(f"### Columns Added ({len(columns_added)})", columns_added)
        if columns_removed:
            st.write(f"### Columns Removed ({len(columns_removed)})", columns_removed)
        if changed is not None and not changed.empty:
            st.write("### Changed Rows", changed)

        if sql_overlay_df is not None and not sql_overlay_df.empty:
            tab1, tab2 = st.tabs(["SQL Changes Table", "SQL Overlay View"])
            with tab1:
                st.write("### SQL Changes Table")
                changes_only = sql_overlay_df[sql_overlay_df['Status'].isin(['Added', 'Removed'])]
                if not changes_only.empty:
                    for _, row in changes_only.iterrows():
                        if row['Status'] == 'Added':
                            st.markdown(f"<span style='color: #006100; background-color: #C6EFCE;'>+ (Line {row['LineNum2']}) {row['Line']}</span>", unsafe_allow_html=True)
                        elif row['Status'] == 'Removed':
                            st.markdown(f"<span style='color: #9C0006; background-color: #FFC7CE;'>- (Line {row['LineNum1']}) {row['Line']}</span>", unsafe_allow_html=True)
                else:
                    st.info("No added or removed lines.")
            with tab2:
                st.write("### SQL Overlay")
                html_lines = []
                for _, row in sql_overlay_df.iterrows():
                    prefix = ''
                    if row['Status'] == 'Added':
                        prefix = f"+ (Line {row['LineNum2']}) "
                        html_lines.append(f"<span style='color: #006100; background-color: #C6EFCE;'>{prefix}{row['Line']}</span>")
                    elif row['Status'] == 'Removed':
                        prefix = f"- (Line {row['LineNum1']}) "
                        html_lines.append(f"<span style='color: #9C0006; background-color: #FFC7CE;'>{prefix}{row['Line']}</span>")
                    else:
                        prefix = f"  (Line {row['LineNum1']}) "
                        html_lines.append(f"<span>{prefix}{row['Line']}</span>")
                st.markdown("<br>".join(html_lines), unsafe_allow_html=True)
    else:
        st.warning("Please upload both CSV files and select at least one key column.")

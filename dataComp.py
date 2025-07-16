import pandas as pd
import argparse
import openpyxl
from openpyxl.styles import PatternFill

def load_csv(file_path):
    return pd.read_csv(file_path)

def get_composite_key(df, key_columns):
    # Strip whitespace and fill NaNs with a placeholder
    return df[key_columns].astype(str).apply(lambda x: x.str.strip()).fillna('NA').agg('-'.join, axis=1)

def check_unique_composite_key(df, composite_key):
    duplicates = composite_key[composite_key.duplicated(keep=False)]
    if not duplicates.empty:
        print("Warning: Duplicate composite keys found:")
        print(df.loc[composite_key.isin(duplicates), :])
        return False
    return True

def convert_numeric_columns(df, numeric_columns):
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='ignore')
    return df

def strip_leading_zeros(df):
    # Apply to all cells: if it's a string of digits, strip leading zeros
    return df.map(lambda x: x.lstrip('0') if isinstance(x, str) and x.isdigit() else x)

def compare_data(file1, file2, key_columns=None, numeric_columns=None):
    df1 = load_csv(file1)
    df2 = load_csv(file2)

    # Strip leading zeros from all columns
    df1 = strip_leading_zeros(df1)
    df2 = strip_leading_zeros(df2)

    # Make all columns strings for consistent comparison
    df1 = df1.astype(str)
    df2 = df2.astype(str)

    if key_columns is None:
        key_columns = df1.columns.tolist()

    df1['composite_key'] = get_composite_key(df1, key_columns)
    df2['composite_key'] = get_composite_key(df2, key_columns)

    # Check uniqueness of composite key
    if not check_unique_composite_key(df1, df1['composite_key']):
        print(f"Duplicate composite keys found in {file1}. Please fix before proceeding.")
        return
    if not check_unique_composite_key(df2, df2['composite_key']):
        print(f"Duplicate composite keys found in {file2}. Please fix before proceeding.")
        return

    df1.set_index('composite_key', inplace=True)
    df2.set_index('composite_key', inplace=True)

    added = df2.loc[~df2.index.isin(df1.index)]
    removed = df1.loc[~df1.index.isin(df2.index)]
    common = df1.loc[df1.index.isin(df2.index)]

    changed = []
    for key in common.index:
        if not df1.loc[key].equals(df2.loc[key]):
            diff = pd.concat([df1.loc[[key]], df2.loc[[key]]], keys=['file1', 'file2'])
            changed.append(diff)
    changed_df = pd.concat(changed) if changed else pd.DataFrame()

    with pd.ExcelWriter('comparison_results.xlsx') as writer:
        added.to_excel(writer, sheet_name='Added')
        removed.to_excel(writer, sheet_name='Removed')
        if not changed_df.empty:
            changed_df.to_excel(writer, sheet_name='Changed')

def compare_sql_files(sql_file1, sql_file2):
    with open(sql_file1, 'r') as f1, open(sql_file2, 'r') as f2:
        sql1_lines = [line.strip() for line in f1.readlines()]
        sql2_lines = [line.strip() for line in f2.readlines()]

    diff_rows = []
    # Track lines by content and index
    sql1_dict = {line: idx+1 for idx, line in enumerate(sql1_lines)}
    sql2_dict = {line: idx+1 for idx, line in enumerate(sql2_lines)}

    all_lines = set(sql1_lines) | set(sql2_lines)
    for line in all_lines:
        in_file1 = line in sql1_dict
        in_file2 = line in sql2_dict
        if in_file1 and not in_file2:
            diff_rows.append({'Line': sql1_dict[line], 'File1': line, 'File2': '', 'Status': 'Removed'})
        elif not in_file1 and in_file2:
            diff_rows.append({'Line': sql2_dict[line], 'File1': '', 'File2': line, 'Status': 'Added'})
        # If line exists in both, skip (or handle as needed)
    return pd.DataFrame(diff_rows)

if __name__ == "__main__":
    print("--- Data Comparison Script ---")
    file1 = input("Enter path to first CSV file: ").strip()
    file2 = input("Enter path to second CSV file: ").strip()
    key_columns = input("Enter comma-separated key columns (leave blank to use all columns): ").strip()
    key_columns = key_columns.split(",") if key_columns else None
    sql1 = input("Enter path to first SQL file (optional, leave blank to skip): ").strip()
    sql2 = input("Enter path to second SQL file (optional, leave blank to skip): ").strip()
    sql1 = sql1 if sql1 else None
    sql2 = sql2 if sql2 else None

    compare_data(file1, file2, key_columns)

    # Optionally compare SQL files and add overlay sheet to Excel
    if sql1 and sql2:
        import difflib
        with open(sql1, 'r') as f1, open(sql2, 'r') as f2:
            sql1_lines = [line.rstrip('\n') for line in f1]
            sql2_lines = [line.rstrip('\n') for line in f2]

        diff = list(difflib.ndiff(sql1_lines, sql2_lines))
        overlay_rows = []
        for line in diff:
            if line.startswith('+ '):
                overlay_rows.append({'Line': line[2:], 'Status': 'Added'})
            elif line.startswith('- '):
                overlay_rows.append({'Line': line[2:], 'Status': 'Removed'})
            elif line.startswith('  '):
                overlay_rows.append({'Line': line[2:], 'Status': 'Unchanged'})
        df_overlay = pd.DataFrame(overlay_rows)
        with pd.ExcelWriter('comparison_results.xlsx', mode='a', engine='openpyxl') as writer:
            df_overlay.to_excel(writer, sheet_name='SQL_Overlay', index=False)

        # Apply coloring
        wb = openpyxl.load_workbook('comparison_results.xlsx')
        ws = wb['SQL_Overlay']
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
            status = row[1].value
            if status == 'Added':
                row[0].font = openpyxl.styles.Font(color='006100')
                row[0].fill = green_fill
            elif status == 'Removed':
                row[0].font = openpyxl.styles.Font(color='9C0006')
                row[0].fill = red_fill
        wb.save('comparison_results.xlsx')
    print("Comparison completed. Results saved to 'comparison_results.xlsx'.")